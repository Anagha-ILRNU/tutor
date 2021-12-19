import softest
import time
from openpyxl import workbook, load_workbook


class Utils:
    # def assertTitleText(self, title):
    #     act_title = self.driver.title
    #     print(act_title)
    #     time.sleep(5)
    #     self.soft_assert(self.assertEqual,act_title, title)
    #     if act_title == title:
    #         print("Test Passed")
    #     else:
    #         print("Test Failed")

    def read_data_from_excel(file_name ,sheet):
        datalist = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column
        for i in range(2, row_ct, + 1):
            row = []
            for j in range(1, col_ct, + 1):
                row.append(sh.cell(row=i, column=j).value)
            datalist.append(row)
        return datalist
