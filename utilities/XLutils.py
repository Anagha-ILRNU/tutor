import openpyxl

def getRowCount(filename, sheetName):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetName]
    return sheet.max_row


def getColumnCount(filename, sheetName):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetName]
    return sheet.max_column


def readData(filename, sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetName]
    return sheet.cell(row=rownum, column=columnno).value


def writeData(filename, sheetName, rownum, columnno, data):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetName]
    sheet.cell(row=rownum, column=columnno).value = data
    workbook.save(filename)